from time import sleep
import selenium
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
import os
import datetime
from openpyxl import Workbook, load_workbook

options = Options()
options.add_experimental_option(
    'prefs',
    {
        # 'profile.managed_default_content_settings.javascript': 2,
        'profile.managed_default_content_settings.images': 2,
        # 'profile.managed_default_content_settings.mixed_script': 2,
        'profile.managed_default_content_settings.media_stream': 2,
        'profile.managed_default_content_settings.stylesheets': 2
    }
)


def web():
    # def clear_name(data):
    #     """Очистка наименования от [лишних значений]"""
    #     end = data.find('[') - 1
    #     return data[0:end]

    dict_product = {}
    ch = webdriver.Chrome(options=options)

    ch.get("https://www.dns-shop.ru/catalog/17a8a01d16404e77/smartfony/")

    for _ in range(100):
        sleep(3)
        products = ch.find_elements_by_class_name('catalog-product')
        for i in products:
            name_product = i.find_element_by_class_name('catalog-product__name').text
            price_product = i.find_element_by_class_name('product-buy__price').text

            try:
                promo = i.find_element_by_class_name('w-product-voblers').text
            except selenium.common.exceptions.NoSuchElementException:
                promo = 'нет акций'
            dict_product[name_product] = [price_product, promo]

        button_next = ch.find_element_by_class_name('pagination-widget__page-link_next')
        if button_next.get_attribute('href') == 'javascript:':
            break
        else:
            button_next.click()

    ch.close()
    return dict_product


def writing_file_excel2(price_list, name_f):
    """ Входные данные - словарь"""
    current_time = datetime.datetime.now().strftime("%d-%m-%y_%H-%M") + '_'
    head, tail = os.path.split(__file__)
    name_f = os.path.normpath(f'{head}/data/{name_f}_{current_time}.xlsx')

    wb = Workbook()
    ws = wb.active
    row = 1
    for key, value_list in price_list.items():
        ws.cell(row=row, column=1, value=key)
        column = 2
        for value in value_list:
            ws.cell(row=row, column=column, value=value)
            column += 1
        row += 1

    wb.save(filename=name_f)


data = web()
writing_file_excel2(data, 'raw_data')
